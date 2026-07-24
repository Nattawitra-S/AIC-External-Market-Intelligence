"""
Pivot Cache Extractor
---------------------
วาง .xlsx ไฟล์ใน /input แล้วรัน script นี้
Output จะออกมาใน /output อัตโนมัติ

Fixes in this version:
- ไม่บังคับว่าต้องมี column ชื่อ Total เท่านั้น
- หา numeric/measure column ให้อัตโนมัติ แล้วสร้าง column มาตรฐานชื่อ Total ให้ใช้ต่อ
- ถ้าเป็น generic/raw data และมีมากกว่า Excel limit จะ split หลาย sheets ให้อัตโนมัติ
- ใช้ csv.writer เพื่อกันปัญหา comma/quote ในข้อมูล

This is the tracked (git-visible) copy of the extractor. Its data
directories (input/output) live under
raw_data/department_of_education/File_extractor 2/, which is gitignored —
so the script location and the data location are deliberately different.
Do not use Path(__file__).parent as the data root here.
"""

import argparse
import csv
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
# This script lives at ETL/tools/extract_fixed.py (tracked). Its data
# directories live at
# raw_data/department_of_education/File_extractor 2/{input,output}
# (gitignored). Resolve the project root from this file's own location
# rather than assuming the script and data share a directory.
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR     = PROJECT_ROOT / "raw_data" / "department_of_education" / "File_extractor 2"
INPUT_DIR    = DATA_DIR / "input"
OUTPUT_DIR   = DATA_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Files processed automatically when no filename argument is supplied.
# Deterministic, matching the canonical local input filenames used by the
# rest of the Education pipeline (ETL/etl_education_v2.py) -- not a glob.
AUTO_INPUT_FILES = [
    "Pivot_Basic_All_web.xlsx",
    "Pivot_Detailed_Latest_web.xlsx",
]

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_MAP = {"x": NS}

# Excel supports 1,048,576 rows including header.
EXCEL_MAX_ROWS = 1_048_576
DATA_ROWS_PER_SHEET = EXCEL_MAX_ROWS - 1

# ── Styles ─────────────────────────────────────────────────────────────────────
FILLS = {
    "blue":   PatternFill("solid", start_color="1F4E79"),
    "green":  PatternFill("solid", start_color="1E4D2B"),
    "purple": PatternFill("solid", start_color="4A235A"),
    "teal":   PatternFill("solid", start_color="0E4D4D"),
    "alt":    PatternFill("solid", start_color="EBF3FF"),
    "alt2":   PatternFill("solid", start_color="EBF3FF"),
}
WH = Font(bold=True, color="FFFFFF", name="Arial", size=10)


def safe_sheet_title(title: str, fallback: str = "Sheet") -> str:
    """Excel sheet names must be <=31 chars and cannot contain []:*?/\\."""
    title = re.sub(r"[\[\]:*?/\\]", " ", str(title)).strip() or fallback
    return title[:31]


# ── Step 1: Extract raw data from pivot cache ──────────────────────────────────
def extract_pivot_cache(xlsx_path: Path) -> tuple[pd.DataFrame, list[str]]:
    print(f"  Reading pivot cache from: {xlsx_path.name}")

    with zipfile.ZipFile(xlsx_path) as z:
        all_files = z.namelist()

        defn_files = sorted([
            f for f in all_files
            if "pivotCacheDefinition" in f and not f.endswith(".rels")
        ])
        rec_files = sorted([
            f for f in all_files
            if "pivotCacheRecords" in f and not f.endswith(".rels")
        ])

        if not defn_files:
            raise ValueError("No pivot cache definition found in this file.")
        if not rec_files:
            raise ValueError("No pivot cache records found in this file.")

        with z.open(defn_files[0]) as f:
            defn = ET.fromstring(f.read())

        fields = defn.findall(".//x:cacheField", NS_MAP)
        field_names = [field.get("name") or f"Field_{i+1}" for i, field in enumerate(fields)]

        field_values: dict[str, list[str]] = {}
        for field in fields:
            name = field.get("name") or ""
            strings = [s.get("v", "") for s in field.findall(".//x:s", NS_MAP)]
            nums = [n.get("v", "") for n in field.findall(".//x:n", NS_MAP)]
            field_values[name] = strings if strings else nums

        print("  Fields found:")
        for name in field_names:
            print(f"    - {name}")

        tag_r = f"{{{NS}}}r"
        tag_x = f"{{{NS}}}x"
        tag_n = f"{{{NS}}}n"
        tag_m = f"{{{NS}}}m"
        tag_s = f"{{{NS}}}s"  # inline string, uncommon but supported

        rows_written = 0
        csv_path = OUTPUT_DIR / f"_raw_{xlsx_path.stem}.csv"

        with z.open(rec_files[0]) as xmlf, open(csv_path, "w", encoding="utf-8", newline="") as csvf:
            writer = csv.writer(csvf)
            writer.writerow(field_names)

            record: list[str] = []
            field_idx = 0

            for event, elem in ET.iterparse(xmlf, events=("start", "end")):
                if event == "start":
                    if elem.tag == tag_r:
                        record = []
                        field_idx = 0

                    elif elem.tag == tag_x:
                        idx = int(elem.get("v", 0))
                        fname = field_names[field_idx] if field_idx < len(field_names) else ""
                        vals = field_values.get(fname, [])
                        record.append(vals[idx] if idx < len(vals) else "")
                        field_idx += 1

                    elif elem.tag == tag_n:
                        record.append(elem.get("v", ""))
                        field_idx += 1

                    elif elem.tag == tag_s:
                        record.append(elem.get("v", ""))
                        field_idx += 1

                    elif elem.tag == tag_m:
                        record.append("")
                        field_idx += 1

                elif event == "end" and elem.tag == tag_r:
                    if len(record) < len(field_names):
                        record.extend([""] * (len(field_names) - len(record)))
                    elif len(record) > len(field_names):
                        record = record[:len(field_names)]

                    writer.writerow(record)
                    rows_written += 1

                    if rows_written % 200_000 == 0:
                        print(f"    {rows_written:,} rows extracted...")
                    elem.clear()

    print(f"  ✓ {rows_written:,} total rows extracted")

    df = pd.read_csv(csv_path, dtype=str, low_memory=False)
    csv_path.unlink(missing_ok=True)

    # Create standard measure column named Total only when a real measure is found.
    value_col = detect_value_column(df)
    if value_col:
        df["Total"] = pd.to_numeric(df[value_col], errors="coerce").fillna(0).astype(int)
        print(f"  ✓ Measure column detected: {value_col} -> Total")
    else:
        print("  ⚠ No numeric measure column detected. Generic raw output will still be created.")

    return df, field_names


def detect_value_column(df: pd.DataFrame) -> str | None:
    """Find the best numeric/measure column. The old script assumed it was always 'Total'."""
    if "Total" in df.columns:
        return "Total"

    preferred_keywords = [
        "total",
        "sum of",
        "count",
        "enrolment",
        "enrollment",
        "grant",
        "grants",
        "data",
        "value",
        "values",
    ]

    candidates: list[tuple[int, int, str]] = []
    for col in df.columns:
        s = pd.to_numeric(df[col], errors="coerce")
        valid_count = int(s.notna().sum())
        if valid_count == 0:
            continue

        lower = str(col).lower()
        keyword_score = sum(1 for kw in preferred_keywords if kw in lower)
        candidates.append((keyword_score, valid_count, col))

    if not candidates:
        return None

    # Prioritise measure-like names, then most numeric values.
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


# ── Step 2: Detect file type and build appropriate output ──────────────────────
def detect_file_type(field_names: list[str], df: pd.DataFrame) -> str:
    """Guess what kind of report this is based on field names/columns."""
    fields_set = set(field_names) | set(df.columns)

    if "Citizenship Country" in fields_set and "Nominated Occupation" in fields_set:
        return "skilled"
    if "Citizenship Country" in fields_set and "Visa Type" in fields_set:
        return "whm"
    if {"Sector", "Nationality"}.issubset(fields_set) or "YTD Enrolments" in " ".join(fields_set):
        return "education"
    return "generic"


# ── Step 3: Build Excel output ─────────────────────────────────────────────────
def write_pivot_sheet(ws, pivot_df: pd.DataFrame, fill, index_cols: int = 1):
    col_names = list(pivot_df.columns)
    for ci, h in enumerate(col_names, 1):
        c = ws.cell(1, ci, h)
        c.font = WH
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    toggle = [False]
    prev = [None]
    for ri, row in enumerate(pivot_df.itertuples(index=False), 2):
        key = row[0]
        if key != prev[0]:
            prev[0] = key
            toggle[0] = not toggle[0]
        for ci, v in enumerate(row, 1):
            if ci <= index_cols:
                cell_value = v
            else:
                try:
                    cell_value = int(v)
                except Exception:
                    cell_value = v
            c = ws.cell(ri, ci, cell_value)
            if toggle[0]:
                c.fill = FILLS["alt"]

    ws.freeze_panes = f"{get_column_letter(index_cols + 1)}2"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_skilled_excel(df: pd.DataFrame, out_path: Path):
    require_columns(df, ["Citizenship Country", "Nominated Occupation", "Financial Year of Visa Grant", "Total"])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Country × Detailed Occupation"
    p = df.groupby(["Citizenship Country", "Nominated Occupation", "Financial Year of Visa Grant"], dropna=False)["Total"].sum().reset_index()
    p = p.pivot_table(index=["Citizenship Country", "Nominated Occupation"], columns="Financial Year of Visa Grant", values="Total", aggfunc="sum", fill_value=0)
    p["Grand Total"] = p.sum(axis=1)
    p = p.sort_values(["Citizenship Country", "Grand Total"], ascending=[True, False]).reset_index()
    write_pivot_sheet(ws1, p, FILLS["blue"], index_cols=2)
    set_widths(ws1, [35, 50] + [11] * (len(p.columns) - 2))
    print("  ✓ Sheet: Country × Detailed Occupation")

    if "Nominated Occupation (Unit Group)" in df.columns:
        ws2 = wb.create_sheet("Country × Unit Group")
        p = df.groupby(["Citizenship Country", "Nominated Occupation (Unit Group)", "Financial Year of Visa Grant"], dropna=False)["Total"].sum().reset_index()
        p = p.pivot_table(index=["Citizenship Country", "Nominated Occupation (Unit Group)"], columns="Financial Year of Visa Grant", values="Total", aggfunc="sum", fill_value=0)
        p["Grand Total"] = p.sum(axis=1)
        p = p.sort_values(["Citizenship Country", "Grand Total"], ascending=[True, False]).reset_index()
        write_pivot_sheet(ws2, p, FILLS["green"], index_cols=2)
        set_widths(ws2, [35, 45] + [11] * (len(p.columns) - 2))
        print("  ✓ Sheet: Country × Unit Group")

    if "Nominated Occupation (Major Group)" in df.columns:
        ws3 = wb.create_sheet("Country × Major Group")
        p = df.groupby(["Citizenship Country", "Nominated Occupation (Major Group)", "Financial Year of Visa Grant"], dropna=False)["Total"].sum().reset_index()
        p = p.pivot_table(index=["Citizenship Country", "Nominated Occupation (Major Group)"], columns="Financial Year of Visa Grant", values="Total", aggfunc="sum", fill_value=0)
        p["Grand Total"] = p.sum(axis=1)
        p = p.sort_values(["Citizenship Country", "Grand Total"], ascending=[True, False]).reset_index()
        write_pivot_sheet(ws3, p, FILLS["teal"], index_cols=2)
        set_widths(ws3, [35, 42] + [11] * (len(p.columns) - 2))
        print("  ✓ Sheet: Country × Major Group")

    ws4 = wb.create_sheet("Country by Year")
    p = df.groupby(["Citizenship Country", "Financial Year of Visa Grant"], dropna=False)["Total"].sum().reset_index()
    p = p.pivot_table(index="Citizenship Country", columns="Financial Year of Visa Grant", values="Total", aggfunc="sum", fill_value=0)
    p["Grand Total"] = p.sum(axis=1)
    p = p.sort_values("Grand Total", ascending=False).reset_index()
    write_pivot_sheet(ws4, p, FILLS["purple"], index_cols=1)
    set_widths(ws4, [35] + [11] * (len(p.columns) - 1))
    print("  ✓ Sheet: Country by Year")

    wb.save(out_path)


def build_whm_excel(df: pd.DataFrame, out_path: Path):
    require_columns(df, ["Citizenship Country", "Financial Year of Visa Grant", "Visa Type", "Total"])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Country by Year"
    p = df.groupby(["Citizenship Country", "Financial Year of Visa Grant"], dropna=False)["Total"].sum().reset_index()
    p = p.pivot_table(index="Citizenship Country", columns="Financial Year of Visa Grant", values="Total", aggfunc="sum", fill_value=0)
    p["Grand Total"] = p.sum(axis=1)
    p = p.sort_values("Grand Total", ascending=False).reset_index()
    write_pivot_sheet(ws1, p, FILLS["blue"], index_cols=1)
    set_widths(ws1, [35] + [11] * (len(p.columns) - 1))
    print("  ✓ Sheet: Country by Year")

    ws2 = wb.create_sheet("Country × Visa Type")
    p = df.groupby(["Citizenship Country", "Visa Type"], dropna=False)["Total"].sum().reset_index()
    p = p.pivot_table(index="Citizenship Country", columns="Visa Type", values="Total", aggfunc="sum", fill_value=0)
    p["Grand Total"] = p.sum(axis=1)
    p = p.sort_values("Grand Total", ascending=False).reset_index()
    write_pivot_sheet(ws2, p, FILLS["teal"], index_cols=1)
    set_widths(ws2, [35] + [20] * (len(p.columns) - 1))
    print("  ✓ Sheet: Country × Visa Type")

    wb.save(out_path)


def build_education_excel(df: pd.DataFrame, out_path: Path):
    """Education/enrolment style report. Creates useful summary sheets where columns exist."""
    if "Total" not in df.columns:
        build_generic_excel(df, out_path)
        return

    wb = Workbook()
    first_sheet_used = False

    def add_summary(sheet_name: str, group_cols: list[str], fill):
        nonlocal first_sheet_used
        if not all(c in df.columns for c in group_cols):
            return

        ws = wb.active if not first_sheet_used else wb.create_sheet(safe_sheet_title(sheet_name))
        ws.title = safe_sheet_title(sheet_name)
        first_sheet_used = True

        p = df.groupby(group_cols, dropna=False)["Total"].sum().reset_index()
        p = p.sort_values("Total", ascending=False)
        write_pivot_sheet(ws, p, fill, index_cols=len(group_cols))
        set_widths(ws, [28] * len(group_cols) + [14])
        print(f"  ✓ Sheet: {sheet_name}")

    add_summary("Sector Summary", ["Sector"], FILLS["blue"])
    add_summary("Nationality Summary", ["Nationality"], FILLS["green"])
    add_summary("State Summary", ["State"], FILLS["teal"])
    add_summary("ProviderType Summary", ["ProviderType"], FILLS["purple"])
    add_summary("Sector by Nationality", ["Sector", "Nationality"], FILLS["blue"])
    add_summary("State by Nationality", ["State", "Nationality"], FILLS["green"])

    if not first_sheet_used:
        # Remove the blank default workbook sheet by replacing with generic raw split output.
        build_generic_excel(df, out_path)
        return

    wb.save(out_path)

    # Also save full raw data as split workbook because the summary workbook will not contain every raw row.
    raw_path = out_path.with_name(out_path.stem + "_raw_split.xlsx")
    build_generic_excel(df, raw_path)
    print(f"  ✓ Full raw data also saved to: {raw_path}")


def require_columns(df: pd.DataFrame, cols: list[str]):
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Available columns: {list(df.columns)}")


def build_generic_excel(df: pd.DataFrame, out_path: Path):
    """Fallback: write all data, split across sheets if it exceeds Excel's row limit."""
    print(f"  Writing generic raw output: {len(df):,} rows")

    wb = Workbook(write_only=True)
    cols = list(df.columns)

    total_rows = len(df)
    sheet_count = max(1, (total_rows + DATA_ROWS_PER_SHEET - 1) // DATA_ROWS_PER_SHEET)

    for sheet_no in range(sheet_count):
        start = sheet_no * DATA_ROWS_PER_SHEET
        end = min(start + DATA_ROWS_PER_SHEET, total_rows)
        ws = wb.create_sheet(safe_sheet_title(f"All Data {sheet_no + 1}"))

        header = []
        for h in cols:
            cell = WriteOnlyCell(ws, value=h)
            cell.font = WH
            cell.fill = FILLS["blue"]
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            header.append(cell)
        ws.append(header)

        for row in df.iloc[start:end].itertuples(index=False, name=None):
            ws.append(list(row))

        print(f"  ✓ Sheet: All Data {sheet_no + 1} ({start + 1:,}–{end:,})")

    wb.save(out_path)
    print("  ✓ Generic raw output saved")


# ── Main ───────────────────────────────────────────────────────────────────────
def process_file(xlsx_path: Path):
    print(f"\n{'=' * 60}")
    print(f"Processing: {xlsx_path.name}")
    print(f"{'=' * 60}")

    df, field_names = extract_pivot_cache(xlsx_path)
    file_type = detect_file_type(field_names, df)
    out_path = OUTPUT_DIR / f"{xlsx_path.stem}_extracted.xlsx"

    print(f"  Detected type: {file_type}")
    print("  Building Excel output...")

    if file_type == "skilled":
        build_skilled_excel(df, out_path)
    elif file_type == "whm":
        build_whm_excel(df, out_path)
    elif file_type == "education":
        build_education_excel(df, out_path)
    else:
        build_generic_excel(df, out_path)

    print("\n  ✅ Done! Output saved to:")
    print(f"     {out_path}")


def resolve_input_path(filename: str) -> Path:
    """
    Resolve `filename` strictly inside INPUT_DIR.

    Rejects absolute paths, path traversal outside INPUT_DIR, non-.xlsx
    files, and files that don't exist. Only a bare filename (optionally with
    subdirectories) relative to INPUT_DIR is accepted.
    """
    raw = Path(filename)

    if raw.is_absolute():
        raise ValueError(f"Absolute paths are not allowed: {filename}")
    if raw.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx files are accepted: {filename}")

    input_dir_resolved = INPUT_DIR.resolve()
    candidate = (INPUT_DIR / raw).resolve()

    if candidate != input_dir_resolved and input_dir_resolved not in candidate.parents:
        raise ValueError(f"Path escapes input directory: {filename}")
    if not candidate.is_file():
        raise FileNotFoundError(f"Input file not found: {candidate}")

    return candidate


def main():
    ap = argparse.ArgumentParser(
        description="Extract a pivot-cache .xlsx workbook (from /input) into a flat/tabular output (in /output).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python3 ETL/tools/extract_fixed.py Pivot_Basic_All_web.xlsx\n"
            "  python3 ETL/tools/extract_fixed.py\n\n"
            "With no filename, processes whichever of these are present in input/:\n"
            + "\n".join(f"  - {name}" for name in AUTO_INPUT_FILES) + "\n\n"
            f"Input folder:  {INPUT_DIR}\n"
            f"Output folder: {OUTPUT_DIR}"
        ),
    )
    ap.add_argument(
        "filename",
        nargs="?",
        help="Optional name of the .xlsx file inside the input/ folder to process (not a full "
             "path). Omit to automatically process Pivot_Basic_All_web.xlsx and "
             "Pivot_Detailed_Latest_web.xlsx, whichever are present.",
    )
    args = ap.parse_args()

    if args.filename:
        # Explicit single-filename mode: byte-for-byte the original,
        # unchanged CLI contract -- resolve exactly this one file, fail
        # immediately (using the exception's own message) if it's missing
        # or invalid. No auto-mode tolerance applies here.
        try:
            xlsx_path = resolve_input_path(args.filename)
        except (ValueError, FileNotFoundError) as e:
            print(f"❌ {e}")
            sys.exit(1)
        input_paths = [xlsx_path]
    else:
        # Auto mode: process whichever of the deterministic Basic/Detailed
        # filenames are present, skip a missing one with a warning, and
        # fail only if neither is present.
        input_paths = []
        missing_files = []
        for filename in AUTO_INPUT_FILES:
            try:
                input_paths.append(resolve_input_path(filename))
            except FileNotFoundError:
                missing_files.append(filename)
            except ValueError as e:
                # Absolute path / traversal / non-.xlsx in a hardcoded
                # deterministic filename would mean real misconfiguration,
                # not "this month's file isn't published yet" -- fatal.
                print(f"❌ {e}")
                sys.exit(1)

        if missing_files:
            print("⚠ Skipping file(s) not currently present in the input folder:")
            for filename in missing_files:
                print(f"   - {filename}")

        if not input_paths:
            print("\n❌ No supported input files were found.")
            print("   Add at least one of these files to the input folder:")
            for filename in AUTO_INPUT_FILES:
                print(f"   - {filename}")
            sys.exit(1)

    for xlsx_path in input_paths:
        try:
            process_file(xlsx_path)
        except Exception as e:
            print(f"\n  ❌ Error processing {xlsx_path.name}: {e}")
            sys.exit(1)

    print(f"\n{'=' * 60}")
    print("All done! Check the /output folder.")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()
